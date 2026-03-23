import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

data = [
    ["STT", "Intent / Khúc mắc", "Câu hỏi mẫu (Khách hàng)", "Câu trả lời mẫu (AI Sommelier)", "Quy tắc áp dụng (Rules) / Notes"],
    [1, "Greeting (Chào hỏi)", "Chào bạn / Alo / Shop ơi", "Chào Quý khách, Nhà Chát Sommelier có thể giúp gì cho Quý khách hôm nay ạ? Quý khách đang tìm vang đỏ, vang trắng hay muốn mua quà tặng ạ?", "Tone giọng thân thiện, gợi mở nhu cầu."],
    [2, "Discovery (Tìm hiểu nhu cầu)", "Mình tính tìm 1 chai tầm 1-2 triệu, mình ăn kèm thịt bò Steak.", "Dạ, với mức giá 1-2 triệu để ăn cùng bò Steak, Quý khách nên chọn các dòng vang đỏ đậm đà giúp tôn vị ngọt của thịt. Quý khách có người quen chuộng vang Pháp hay Ý không ạ?", "Lọc ngân sách và phân tích món ăn."],
    [3, "Recommendation (Đề xuất)", "Gợi ý cho mình loại vang Pháp nhé.", "Nhà Chát xin gợi ý 3 chai vang Pháp tuyệt hảo:\n1. [Chai đắt nhất] - Giá: 1,800,000 VND. ✨ Ưu điểm: Đậm đà, tannin mượt, hợp Steak.\n2. [Chai giữa] - Giá: 1,500,000 VND. ✨ Ưu điểm: ...\n3. [Chai rẻ nhất] - Giá: 1,200,000 VND. ✨ Ưu điểm: ...", "LUÔN XẾP GIÁ TỪ CAO XUỐNG THẤP. Mộc mạc, không dùng thuật ngữ khó hiểu."],
    [4, "Blacklist (Xử lý sản phẩm lỗi)", "Bên bạn có hộp sơn mài 2 nắp kính không?", "Dạ hiện tại Nhà Chát không cung cấp mẫu hộp sơn mài 2 nắp kính ạ. Quý khách có muốn tham khảo các mẫu hộp cao cấp nguyên bản từ nhà sản xuất không ạ?", "TUYỆT ĐỐI KHÔNG gợi ý hộp sơn mài 2 nắp kính."],
    [5, "Closing & VAT (Chốt Sales)", "Ok, lấy cho mình chai đầu. Cho mình xuất hóa đơn.", "Dạ vâng ạ. Nhà Chát có xuất hóa đơn với GTGT 10% cho tất cả các sản phẩm rượu. Quý khách hàng cần cung cấp đầy đủ thông tin xuất hóa đơn khi lên đơn hàng qua web hoặc gửi Admin báo qua đây ạ.", "LUÔN nhắc xuất GTGT 10% theo nguyên văn nhắc nhở."],
    [6, "Chống đi lạc đề (Out of bounds)", "Cách phân biệt đồng hồ Rolex thật giả?", "Dạ, em là Sommelier của Nhà Chát, chuyên hỗ trợ tư vấn các dòng rượu vang cao cấp. Em rất tiếc chưa nắm rõ về đồng hồ ạ. Quý khách có muốn em gợi ý một chai vang đang bán chạy không ạ?", "Lịch sự đưa câu chuyện về lại sản phẩm rượu vang."],
    [7, "Trường hợp thêm 1 (User tự điền)", "", "", "Ghi chú rule vào đây.."],
    [8, "Trường hợp thêm 2 (User tự điền)", "", "", ""],
    [9, "Trường hợp thêm 3 (User tự điền)", "", "", ""],
    [10, "Trường hợp thêm 4 (User tự điền)", "", "", ""],
    [11, "Trường hợp thêm 5 (User tự điền)", "", "", ""]
]

df = pd.DataFrame(data[1:], columns=data[0])
# create artifacts dir path
artifact_dir = r"C:\Users\Admin\.gemini\antigravity\brain\efe18f84-1df9-48e1-8f7e-db17a52b4dba"
output_path = os.path.join(artifact_dir, "Kich_ban_AI_Sommelier_V1.xlsx")
workspace_path = r"c:\Users\Admin\Desktop\Code\AI-Sommerlier\Kich_ban_AI_Sommelier_V1.xlsx"

for path in [output_path, workspace_path]:
    df.to_excel(path, index=False)
    # Format using openpyxl
    wb = load_workbook(path)
    ws = wb.active

    # Header formatting
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='1F497D')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 30

    # Alignment and wrapping for data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='center', vertical='top')

    wb.save(path)
    print(f"Saved to {path}")
