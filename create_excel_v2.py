import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

data = [
    ["Giai đoạn (BA Phase)", "Intent / Khúc mắc", "Câu hỏi mẫu (Khách hàng)", "Câu trả lời mẫu (AI Sommelier)", "Quy tắc áp dụng (Rules) / Notes"],
    ["01_Elicitation", "Greeting (Chào hỏi)", "Chào bạn / Shop ơi", "Chào Quý khách, Nhà Chát Sommelier có thể giúp gì cho Quý khách hôm nay ạ?", "Tone giọng thân thiện, chuyên nghiệp."],
    ["01_Elicitation", "Discovery (Tìm hiểu nhu cầu)", "Mình tìm vang uống với bò steak", "Dạ, ăn cùng thịt bò Steak Quý khách nên chọn vang đỏ đậm đà (Full-body) như Cabernet Sauvignon. Quý khách dự kiến mức ngân sách khoảng bao nhiêu ạ?", "Khai thác tối đa 3 yếu tố: Món ăn, Ngân sách, Sở thích cá nhân."],
    ["02_Analysis_Design", "Recommendation (Đề xuất)", "Gợi ý cho mình vang Pháp tầm 1-2 triệu", "Nhà Chát xin đề xuất 3 dòng vang Pháp tuyệt hảo:\n1. [Chai cao nhất] - 1,800K\n2. [Chai giữa] - 1,500K\n3. [Chai thấp nhất] - 1,200K", "LUÔN XẾP GIÁ TỪ CAO XUỐNG THẤP. Giới thiệu Tasting notes dễ hiểu, mộc mạc."],
    ["02_Analysis_Design", "Food Pairing (Phân tích món)", "Vang trắng Mastroberardino uống với hải sản được không?", "Dạ hoàn toàn phù hợp ạ. Vang trắng Ý với độ chua thanh mát sẽ làm nổi bật vị tươi ngon của hải sản, đặc biệt là tôm hay hàu sống.", "Sử dụng logic Food Pairing chuẩn Sommelier."],
    ["06_Strategic_Value", "Upsell / Cross-sell", "Ok, mình lấy chai số 1.", "Dạ vâng. Để thưởng thức trọn vẹn hương vị, Quý khách có cần dùng thêm ly vang pha lê hoặc decanter không ạ? Bên em đang có ưu đãi cho phụ kiện ạ.", "Gợi ý nhẹ nhàng, tăng trưởng doanh thu (Chiến lược kinh doanh)."],
    ["07_Verification_Security", "Age Verification", "Mua 1 chai giao luôn nhé.", "Dạ vâng, để tuân thủ quy định pháp luật về rượu bia, Quý khách vui lòng xác nhận mình đã đủ 18 tuổi chưa ạ?", "Bắt buộc kiểm duyệt tuổi trước khi tạo đơn hàng (Compliance)."],
    ["07_Verification_Security", "Blacklist / Lỗi", "Cho mình 1 hộp sơn mài 2 nắp kính.", "Dạ hiện tại Nhà Chát không cung cấp mẫu hộp sơn mài 2 nắp kính ạ. Quý khách tham khảo mẫu hộp giấy cao cấp khác nhé.", "TUYỆT ĐỐI KHÔNG bán/gợi ý hộp sơn mài 2 nắp kính (Security/Risk)."],
    ["07_Verification_Security", "Out of Bounds (Lạc đề)", "Giá vàng hôm nay bao nhiêu?", "Dạ, em là Sommelier của Nhà Chát chuyên hỗ trợ về rượu vang. Quý khách có muốn tham khảo các mẫu rượu vang mới về không ạ?", "Kiểm soát an ninh prompt: Lịch sự đưa câu chuyện về lại rượu vang."],
    ["03_Documentation", "Closing & VAT", "Cho mình lấy chai đó, xuất hóa đơn công ty.", "Dạ vâng ạ. Nhà Chát có xuất hóa đơn với GTGT 10% cho tất cả sản phẩm. Quý khách cung cấp thông tin xuất hóa đơn cho em nhé.", "Tài liệu hóa: LUÔN nhắc xuất GTGT 10% nguyên văn."],
    ["08_General_Handoff", "Human Handoff (Gặp NV)", "Mình muốn hỏi chính sách đại lý, gọi quản lý ra đây.", "Dạ, yêu cầu này em xin phép chuyển thông tin đến nhân viên CSKH hoặc Quản lý để hỗ trợ Quý khách chi tiết hơn ạ.", "Báo CSKH / Quản lý ngay khi khách có yêu cầu lớn hoặc vượt thẩm quyền AI."],
    ["...", "Trường hợp thêm 1 (Dành cho User)", "", "", ""],
    ["...", "Trường hợp thêm 2 (Dành cho User)", "", "", ""]
]

df = pd.DataFrame(data[1:], columns=data[0])

artifact_dir = r"C:\Users\Admin\.gemini\antigravity\brain\efe18f84-1df9-48e1-8f7e-db17a52b4dba"
output_path = os.path.join(artifact_dir, "Kich_ban_AI_Sommelier_V2.xlsx")
workspace_path = r"c:\Users\Admin\Desktop\Code\AI-Sommerlier\Kich_ban_AI_Sommelier_V2.xlsx"

for path in [output_path, workspace_path]:
    df.to_excel(path, index=False)
    # Format using openpyxl
    wb = load_workbook(path)
    ws = wb.active

    # Header formatting
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='1F497D')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 35

    # Alignment and wrapping for data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.column == 1:
                col_val = cell.value
                # Color code based on phase
                if col_val:
                    if "01_" in col_val:
                        cell.fill = PatternFill('solid', start_color='DCE6F1')
                    elif "02_" in col_val:
                        cell.fill = PatternFill('solid', start_color='F2DCDB')
                    elif "03_" in col_val:
                        cell.fill = PatternFill('solid', start_color='EBF1DE')
                    elif "06_" in col_val:
                        cell.fill = PatternFill('solid', start_color='E4DFEC')
                    elif "07_" in col_val:
                        cell.fill = PatternFill('solid', start_color='FDE9D9')
                    elif "08_" in col_val:
                        cell.fill = PatternFill('solid', start_color='F2F2F2')

    wb.save(path)
    print(f"Saved V2 to {path}")
